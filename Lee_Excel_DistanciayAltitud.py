from openpyxl import load_workbook  #Librería para trabajar con archivos Excel

FILE_PATH = 'Excel/prueba2.xlsx'  #Ruta donde se encuentra la hoja de cálculo con respecto a la ruta donde se encuentra este archivo python
SHEET = 'tmpke4yap'  #Es el nombre de la hoja de cálculo con la que se va a trabajar

workbook = load_workbook(FILE_PATH, read_only=True)  #Solo se va a usar el modo lectura
sheet = workbook[SHEET]



def calcula_distancia():
	'''Va iterando cada fila de la columna 0 (solo existe una columna). La última fila del fichero Excel, que es la que interesa porque contiene 
	la distancia total del trayecto, la guardamos en una variable tipo string'''
	for row in sheet.iter_rows():
		ultima_fila=row[0].value  

	'''La variable ultima_fila nos proporciona el valor de la distancia del trayecto real, un valor de la altura sobre el nivel del mar y
	un valor de la altura de la obstrucción con el siguiente formato de string: "distancia(km),altura_terreno(m),altura_obstruccion(m)"'''
	#Se va a separar cada elemento del string y se van a guardar los 3 parámetros en una tabla
	tabla_parametros=ultima_fila.split(',')
	#Se obtienen los parámetros necesarios para nuestro estudio
	distancia=tabla_parametros[0]
	#Se pasa de tipo string a tipo float
	distancia=float(distancia)

	return distancia


def calcula_altura():

	altura_nivel_mar=[] #Tabla que almacena la altura del terreno sobre el nivel del mar para un determinado número de muestras
	i=0
	for row in sheet.iter_rows():
		fila=row[0].value  
		tabla=fila.split(',')
		altura_nivel_mar.insert(i,tabla[1])
		i+=1


	#Se elimina el primer elemento de la tabla porque contiene "Terrain height(m)", que no interesa para el cálculo
	altura_nivel_mar.pop(0)
	sumatorio_altura_nivel_mar=0.0 #Almacena el resultado de la suma de todas las muestras de la altura del terreno sobre el nivel del mar en metros
	i=0
	while i<len(altura_nivel_mar):
		#Se pasa de string a float para poder operar
		altura_nivel_mar[i]=float(altura_nivel_mar[i])
		sumatorio_altura_nivel_mar=altura_nivel_mar[i]+sumatorio_altura_nivel_mar
		i+=1

	#Se calcula la altura media del terreno sobre el nivel del mar en metros dividiendo el sumatorio entre el número de muestras
	altura_media_nivel_mar=sumatorio_altura_nivel_mar/len(altura_nivel_mar)

	return altura_media_nivel_mar


