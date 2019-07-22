from openpyxl import load_workbook  #Librería para trabajar con archivos Excel

FILE_PATH = 'Excel/prueba_latitud.xlsx'  #Ruta donde se encuentra la hoja de cálculo con respecto a la ruta donde se encuentra este archivo python
SHEET = 'tmp67g327'  #Es el nombre de la hoja de cálculo con la que se va a trabajar

workbook = load_workbook(FILE_PATH, read_only=True)  #Solo se va a usar el modo lectura
sheet = workbook[SHEET]



def calcula_latitud():
	'''Va iterando cada fila de la columna 1 (que es la que nos interesa). La última fila del fichero Excel, que es la que interesa porque contiene 
	la latitud de una de las estaciones terrenas, la guardamos en una variable tipo string'''
	for row in sheet.iter_rows():
		ultima_fila=row[1].value  

	'''La variable ultima_fila nos proporciona la latitud de una de las estaciones terrenas 
	con el siguiente formato de string: "grados:minutos:segundos"'''
	#Se va a separar cada elemento del string y se van a guardar los 3 parámetros en una tabla
	tabla_parametros=ultima_fila.split(':')
	#Se obtienen el parámetro necesario para nuestro estudio
	latitud=tabla_parametros[0]
	#Se pasa de tipo string a tipo float
	latitud=float(latitud)

	return latitud
